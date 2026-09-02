import os

from openpyxl import load_workbook

# Prevent imports from requiring a real API key during testing.
os.environ.setdefault("ANTHROPIC_API_KEY", "test-key")

from agents import research_agent
from tools.excel_exporter import export_to_excel


def make_sample_paper():
    """
    Return one simulated structured PubMed record.
    """

    return {
        "pmid": "12345678",
        "doi": "10.1000/example",
        "pubmed_url": (
            "https://pubmed.ncbi.nlm.nih.gov/12345678/"
        ),
        "title": "Example oncology research paper",
        "authors": "Researcher A, Researcher B",
        "journal": "Example Oncology Journal",
        "year": "2026",
        "abstract": "This is a simulated abstract.",
        "summary": {
            "study_design": "Simulated cohort study.",
            "key_findings": "Simulated findings.",
            "clinical_significance": (
                "Manual interpretation required."
            ),
            "limitations": "Simulated record.",
            "keywords": [
                "oncology",
                "radiotherapy",
            ],
        },
        "verification_status": "Simulated test record",
    }


def test_research_generates_each_output_once(monkeypatch):
    """
    Confirm that each exporter is called exactly once.
    """

    papers = [make_sample_paper()]
    calls = {
        "excel": 0,
        "detailed": 0,
        "concise": 0,
    }

    def fake_search_pubmed(query):
        assert query == "test oncology topic"
        return papers

    def fake_excel_exporter(received_papers):
        assert received_papers == papers
        calls["excel"] += 1

    def fake_detailed_exporter(query, received_papers):
        assert query == "test oncology topic"
        assert received_papers == papers
        calls["detailed"] += 1
        return "outputs/detailed.docx"

    def fake_concise_exporter(query, received_papers):
        assert query == "test oncology topic"
        assert received_papers == papers
        calls["concise"] += 1
        return "outputs/concise.docx"

    monkeypatch.setattr(
        research_agent,
        "search_pubmed",
        fake_search_pubmed,
    )

    monkeypatch.setattr(
        research_agent,
        "export_to_excel",
        fake_excel_exporter,
    )

    monkeypatch.setattr(
        research_agent,
        "export_pubmed_report",
        fake_detailed_exporter,
    )

    monkeypatch.setattr(
        research_agent,
        "export_concise_summary",
        fake_concise_exporter,
    )

    result = research_agent.research(
        "test oncology topic"
    )

    assert calls == {
        "excel": 1,
        "detailed": 1,
        "concise": 1,
    }

    assert result["papers"] == papers
    assert result["detailed_report"] == (
        "outputs/detailed.docx"
    )
    assert result["concise_summary"] == (
        "outputs/concise.docx"
    )
    assert result["excel_file"] == (
        "research_results.xlsx"
    )


def test_excel_contains_traceability_columns(tmp_path):
    """
    Confirm that the Excel evidence table preserves source data.
    """

    output_file = tmp_path / "test_evidence_table.xlsx"

    export_to_excel(
        [make_sample_paper()],
        filename=str(output_file),
    )

    workbook = load_workbook(output_file)
    sheet = workbook["Verified Evidence Table"]

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    assert headers == [
        "PMID",
        "DOI",
        "PubMed URL",
        "Title",
        "Authors",
        "Journal",
        "Year",
        "Abstract",
        "Study Design",
        "Key Findings",
        "Clinical Significance",
        "Limitations",
        "Keywords",
        "Verification Status",
    ]

    assert sheet.max_row == 2
    assert sheet["A2"].value == "12345678"
    assert sheet["B2"].value == "10.1000/example"
    assert sheet["C2"].value == (
        "https://pubmed.ncbi.nlm.nih.gov/12345678/"
    )
    assert sheet["C2"].hyperlink is not None
    assert sheet["H2"].value == (
        "This is a simulated abstract."
    )
    assert sheet["N2"].value == (
        "Simulated test record"
    )