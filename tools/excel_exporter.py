from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_to_excel(papers, filename="research_results.xlsx"):
    """
    Export PubMed papers to a professionally formatted Excel evidence table.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verified Evidence Table"

    headers = [
        "PMID",
        "DOI",
        "PubMed URL",
        "Title",
        "Authors",
        "Journal",
        "Year",
        "Abstract",
        "Study Design",
        "Key Findings",
        "Clinical Significance",
        "Limitations",
        "Keywords",
        "Verification Status",
    ]

    sheet.append(headers)

    for paper in papers:
        summary = paper.get("summary", {})

        if not isinstance(summary, dict):
            summary = {}

        keywords = summary.get("keywords", [])

        if isinstance(keywords, list):
            keywords = ", ".join(str(keyword) for keyword in keywords)
        else:
            keywords = str(keywords or "")

        sheet.append([
            paper.get("pmid", ""),
            paper.get("doi", ""),
            paper.get("pubmed_url", ""),
            paper.get("title", ""),
            paper.get("authors", ""),
            paper.get("journal", ""),
            paper.get("year", ""),
            paper.get("abstract", ""),
            summary.get("study_design", ""),
            summary.get("key_findings", ""),
            summary.get("clinical_significance", ""),
            summary.get("limitations", ""),
            keywords,
            paper.get("verification_status", "Needs manual verification"),
        ])

    # Header formatting
    header_fill = PatternFill("solid", fgColor="0B3B60")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[1].height = 35
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    # Column widths
    column_widths = {
        "A": 14,
        "B": 22,
        "C": 25,
        "D": 45,
        "E": 35,
        "F": 25,
        "G": 10,
        "H": 65,
        "I": 25,
        "J": 55,
        "K": 55,
        "L": 45,
        "M": 35,
        "N": 28,
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    # Format data rows
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row[0].number_format = "@"

        pubmed_url_cell = row[2]

        if pubmed_url_cell.value:
            pubmed_url_cell.hyperlink = pubmed_url_cell.value
            pubmed_url_cell.style = "Hyperlink"

    # Add an Excel table when records are available
    if papers:
        table = Table(
            displayName="PubMedEvidenceTable",
            ref=f"A1:N{sheet.max_row}",
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        sheet.add_table(table)

    workbook.save(filename)

    print(f"\nExcel evidence table saved as '{filename}'")